import pandas as pd
import openpyxl as xl
from datetime import datetime
import os

class MaterialsFile:
    def __init__(self, file_name = None, directory = os.path.join(os.path.expanduser("~"), "Downloads")):
        self.file_name = file_name
        self.directory = directory

    def open_file(self, df_file=None):
        if not df_file:
            if not self.file_name:
                raise ValueError("No file name has been provided. Please check your file_name arg and try again.")
            df_file = self.file_name

        file_path = os.path.join(self.directory, df_file)

        # Checking file extension and loading file
        file_extension = os.path.splitext(file_path)[1]
        try:
            if file_extension == ".csv":
                file = pd.read_csv(file_path)
            elif file_extension in (".xlsx", ".xls"):
                file = pd.read_excel(file_path)
            else:
                raise ValueError("Unsupported file format.")
            print("File loaded successfully!")
            return file
        except FileNotFoundError:
            print(f"That file doesn't exist in {self.directory}. Please try again.")
        except (pd.errors.EmptyDataError, pd.errors.ParserError):
            print("There was an issue with the file content. Please check the file.")
        except Exception as e:
            print(f"An error occurred: {e}. Please try again.")
        return None

    def save_file(self, df, headers, filename):
        # Create a new workbook and select the active worksheet
        wb = xl.Workbook()
        ws = wb.active

        # Set headers in the first row of the worksheet
        if self.file_name:
            for col_index, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col_index).value = header

            # Insert data into the Excel worksheet and format dates
            for row_index, (_, row) in enumerate(df.iterrows(), start=2):  # start=2 to account for headers
                for col_index, header in enumerate(headers, start=1):
                    cell = ws.cell(row=row_index, column=col_index)
                    cell.value = row[header]
                    # Check if the column header is 'TXN_DATE' and format if it's a Timestamp
                    if header == "TXN_DATE" and isinstance(row[header], pd.Timestamp):
                        cell.number_format = 'M/D/YYYY HH:MM'

        # Construct the full path and save the workbook
        full_path = os.path.join(self.directory, filename)
        wb.save(full_path)
        print(f"File saved successfully: {filename}")

    def aggregate(self):
        file = self.open_file()
        if "Production" in file:
            # Process Production data
            file['TXN_DATE'] = pd.to_datetime(file['TXN_DATE'], format='%m/%d/%Y %I:%M:%S %p')
            headers = ["PART_NBR", "BIN_ID", "TXN_QTY", "USER NAME", "TXN_DATE", "SUB CODE"]
            filtered_df = file[file["APPLICATION"] == "PICKING"].copy()
            filtered_df["SERIAL"] = filtered_df.apply(lambda row: f"{row['PART_NBR']}{row['BIN_ID']}{row['USER NAME']}{row['TXN_DATE'].strftime('%Y%m%d%H%M')}{row['SUB CODE']}", axis=1)
            grouped_df = filtered_df.groupby("SERIAL").agg({
                "PART_NBR": "first",
                "BIN_ID": "first",
                "TXN_QTY": "sum",
                "USER_NAME": "first",
                "TXN_DATE": "first",
                "SUB_CODE": "first"
            }).reset_index()
            self.save_file(grouped_df, headers, "Sorted Production.xlsx")

        elif "Bin" in file:
            # Process Bin data
            file['COUNT_DATE'] = pd.to_datetime(file['COUNT_DATE'], format='%m/%d/%Y %I:%M:%S %p')
            headers = ["FACILITY_ID", "BIN_SOURCE", "BUILDING", "BIN_ID", "PART_NBR", "PART_DESC", "SYSTEM_QTY", "COUNT_QTY", "DELTA", "COUNT_DATE", "COUNTED_BY"]
            file['COUNT_DAY'] = file['COUNT_DATE'].dt.date
            file['SERIAL'] = file.apply(lambda row: f"{row['FACILITY_ID']}{row['BIN_SOURCE']}{row['BUILDING']}{row['BIN_ID']}{row['PART_NBR']}{row['SYSTEM_QTY']}{row['COUNT_DATE'].strftime('%Y%m%d%H%M')}{row['COUNTED_BY']}", axis=1)
            grouped_df = file.sort_values('COUNT_DATE').groupby(['SERIAL', 'COUNT_DAY']).agg({
                "FACILITY_ID": "last", 
                "BIN_SOURCE": "last", 
                "BUILDING": "last", 
                "BIN_ID": "last", 
                "PART_NBR": "last", 
                "PART_DESC": "last", 
                "SYSTEM_QTY": "last", 
                "COUNT_QTY": "last", 
                "DELTA": "last", 
                "COUNT_DATE": "last", 
                "COUNTED_BY": "last"
            }).reset_index()
            self.save_file(grouped_df, headers, "Sorted Bin Counts.xlsx")

    def ci_setup(self):
        os.chdir(self.directory)
        #Get rid of nearly identical file names to avoid errors
        for file in os.listdir("."):
            if "ci_reorder" in file & "_all" not in file:
                new_name = "min_max"
                os.rename(file, new_name)
        #Gather up file names
        today = datetime.now().date()
        dfs = []

        #Iterate through files and instantiate them as dataframes
        for file in os.listdir():
            file_path = os.path.join(self.directory, file)
            last_modified_date = datetime.fromtimestamp(os.path.getmtime(file_path)).date()
            if last_modified_date == today:
                if "min_max" in file:
                    min_max = self.open_file(file)
                    dfs.append(min_max)
                
                elif "ci_reorder" in file:
                    ci_report = self.open_file(file)
                    dfs.append(ci_report)
                
                elif "ci_shortage" in file:
                    shortage_report = self.open_file(file)
                    dfs.append(shortage_report)
                
                elif "CTB_" in file:
                    ctb_report = self.open_file(file)
                    dfs.append(ctb_report)
                
                elif "OHB_report" in file:
                    ohb_report = self.open_file(file)
                    dfs.append(ohb_report)
                
                elif "Open PO" in file:
                    open_po_report = self.open_file(file)
                    dfs.append(open_po_report)
                
                elif "Production Report" in file:
                    recent_production = self.open_file(file)
                    dfs.append(recent_production)

